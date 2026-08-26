"""
exporters.py — CSV and Excel export for timetables and master data.

The app could print and publish HTML but had no way to get data OUT into a
spreadsheet, which is what schools actually pass around. This covers:

  * timetable by class, by teacher, by room
  * lesson (assignment) list
  * teacher / class / room / subject lists

CSV is written with a UTF-8 BOM and ';' separator, because that is what Excel on a
Turkish Windows install opens correctly by double-click — a plain comma-separated
UTF-8 file shows up as one mangled column, which users read as "the export is
broken".

Excel output needs openpyxl. When it is missing the caller is told so rather than
silently getting a CSV with an .xlsx extension, which Excel refuses to open.
"""
import csv
import os
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:      # pragma: no cover - depends on the install
    HAS_XLSX = False


CSV_DELIMITER = ";"
CSV_ENCODING = "utf-8-sig"   # the BOM is what makes Excel detect UTF-8


# ── Shared helpers ───────────────────────────────────────────────────

def _norm(value) -> str:
    return " ".join(str(value or "").split()).strip()


def grid_dimensions(data_store: dict):
    settings = data_store.get("settings", {}) or {}
    days = settings.get("days")
    if not days:
        count = int(settings.get("day_count", data_store.get("gun_sayisi", 5)))
        days = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma",
                "Cumartesi", "Pazar"][:count]
    periods = int(settings.get("periods", data_store.get("ders_saati", 8))) or 8
    return list(days), periods


def period_labels(data_store: dict, periods: int):
    """Column headers, using the bell schedule when the school has defined one."""
    bells = data_store.get("zil_programi") or data_store.get("bell_times") or {}
    labels = []
    for p in range(periods):
        entry = None
        if isinstance(bells, dict):
            entry = bells.get(str(p)) or bells.get(p)
        elif isinstance(bells, list) and p < len(bells):
            entry = bells[p]
        if isinstance(entry, dict):
            start = entry.get("start") or entry.get("baslangic") or ""
            end = entry.get("end") or entry.get("bitis") or ""
            if start and end:
                labels.append(f"{p + 1}. Ders\n{start}-{end}")
                continue
        labels.append(f"{p + 1}. Ders")
    return labels


def _placement_index(data_store: dict):
    """(entity, day, period) -> placement, for each of the three views."""
    by_class = {}
    by_teacher = {}
    by_room = {}
    for p in data_store.get("grid_placements", []) or []:
        if not isinstance(p, dict):
            continue
        day = int(p.get("day", p.get("col", 0)) or 0)
        start = int(p.get("period", p.get("row", 0)) or 0)
        span = max(1, int(p.get("duration", 1) or 1))
        cls = _norm(p.get("class_name") or p.get("class"))
        teacher = _norm(p.get("teacher_name") or p.get("teacher"))
        room = _norm(p.get("room_name") or p.get("room") or p.get("derslik"))
        for off in range(span):
            slot = (day, start + off)
            if cls:
                by_class[(cls, *slot)] = p
            if teacher:
                by_teacher[(teacher, *slot)] = p
            if room:
                by_room[(room, *slot)] = p
    return by_class, by_teacher, by_room


def _cell_text(placement: dict, view: str) -> str:
    if not placement:
        return ""
    subject = _norm(placement.get("subject_name") or placement.get("subject"))
    teacher = _norm(placement.get("teacher_name") or placement.get("teacher"))
    cls = _norm(placement.get("class_name") or placement.get("class"))
    room = _norm(placement.get("room_name") or placement.get("room") or placement.get("derslik"))

    if view == "classes":
        parts = [subject, teacher]
    elif view == "teachers":
        parts = [subject, cls]
    else:                      # rooms
        parts = [subject, cls, teacher]
    if room and view != "rooms":
        parts.append(room)
    return "\n".join(x for x in parts if x)


def build_timetable_table(data_store: dict, view: str = "classes"):
    """(headers, rows) for a printable timetable grid.

    view: "classes" | "teachers" | "rooms"
    """
    days, periods = grid_dimensions(data_store)
    labels = period_labels(data_store, periods)
    by_class, by_teacher, by_room = _placement_index(data_store)

    if view == "teachers":
        entities = [_norm(t.get("ad") or t.get("name"))
                    for t in data_store.get("ogretmenler", []) or [] if isinstance(t, dict)]
        index, first_col = by_teacher, "Öğretmen"
    elif view == "rooms":
        entities = [_norm(r.get("ad") or r.get("name"))
                    for r in data_store.get("derslikler", []) or [] if isinstance(r, dict)]
        index, first_col = by_room, "Derslik"
    else:
        entities = [_norm(c.get("ad") or c.get("name"))
                    for c in data_store.get("siniflar", []) or [] if isinstance(c, dict)]
        index, first_col = by_class, "Sınıf"

    entities = [e for e in entities if e]

    headers = [first_col]
    for day in days:
        for lbl in labels:
            headers.append(f"{day} {lbl.splitlines()[0]}")

    rows = []
    for entity in entities:
        row = [entity]
        for d in range(len(days)):
            for p in range(periods):
                row.append(_cell_text(index.get((entity, d, p)), view))
        rows.append(row)
    return headers, rows


def build_lesson_list(data_store: dict):
    headers = ["Sınıf", "Ders", "Öğretmen", "Derslik", "Haftalık Saat", "Dağılım"]
    rows = []
    for a in data_store.get("atamalar", []) or []:
        if not isinstance(a, dict):
            continue
        rows.append([
            _norm(a.get("class") or a.get("sinif") or a.get("class_name")),
            _norm(a.get("subject") or a.get("ders")),
            _norm(a.get("teacher") or a.get("ogretmen") or a.get("teacher_name")),
            _norm(a.get("room") or a.get("derslik")),
            a.get("duration") or a.get("saat") or "",
            _norm(a.get("type") or a.get("dagilim")),
        ])
    return headers, rows


def build_entity_list(data_store: dict, kind: str):
    """kind: 'ogretmenler' | 'siniflar' | 'derslikler' | 'dersler'"""
    specs = {
        "ogretmenler": (["Ad", "Kısa Ad", "Branş", "Günlük Max", "Haftalık Max"],
                        ["ad", "kisa", "brans", "max_gunluk", "max_haftalik"]),
        "siniflar":    (["Ad", "Kısa Ad", "Mevcut", "Sınıf Öğretmeni"],
                        ["ad", "kisa", "mevcut", "sinif_ogretmeni"]),
        "derslikler":  (["Ad", "Kısa Ad", "Kapasite", "Bina", "Tür"],
                        ["ad", "kisa", "kapasite", "bina", "tur"]),
        "dersler":     (["Ad", "Kısa Ad", "Renk"], ["ad", "kisa", "renk"]),
    }
    headers, fields = specs.get(kind, (["Ad"], ["ad"]))
    rows = []
    for item in data_store.get(kind, []) or []:
        if isinstance(item, dict):
            rows.append([_norm(item.get(f, "")) for f in fields])
    return headers, rows


# ── Writers ──────────────────────────────────────────────────────────

def write_csv(path: str, headers, rows) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with open(path, "w", encoding=CSV_ENCODING, newline="") as f:
        writer = csv.writer(f, delimiter=CSV_DELIMITER, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in rows:
            # Newlines inside a cell survive the quoting, but Excel renders them only
            # with wrap-on; flatten to " / " so the CSV stays readable either way.
            writer.writerow([str(c).replace("\n", " / ") if c else "" for c in row])
    return path


def write_xlsx(path: str, sheets) -> str:
    """sheets: [(sheet_name, headers, rows), ...]"""
    if not HAS_XLSX:
        raise RuntimeError(
            "Excel çıktısı için 'openpyxl' paketi gerekli. "
            "CSV olarak kaydedebilir veya 'pip install openpyxl' çalıştırabilirsiniz.")

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    body_font = Font(size=9)
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for name, headers, rows in sheets:
        # Excel rejects >31 chars and the characters below in a sheet name.
        safe = str(name)[:31]
        for ch in "[]:*?/\\":
            safe = safe.replace(ch, "-")
        ws = wb.create_sheet(safe or "Sayfa")

        ws.append(list(headers))
        for cell in ws[1]:
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = centered
            cell.border = border

        for row in rows:
            ws.append(list(row))

        widths = [len(str(h)) for h in headers]
        for row in rows:
            for i, value in enumerate(row):
                if i < len(widths):
                    longest = max((len(part) for part in str(value or "").split("\n")), default=0)
                    widths[i] = max(widths[i], longest)
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 9), 34)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.border = border
                cell.alignment = left if cell.column == 1 else centered

        ws.freeze_panes = "B2"          # keep the entity column and header visible
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    wb.save(path)
    return path


# ── High-level entry points ──────────────────────────────────────────

EXPORT_KINDS = [
    ("timetable_classes", "Ders Programı — Sınıflara Göre"),
    ("timetable_teachers", "Ders Programı — Öğretmenlere Göre"),
    ("timetable_rooms", "Ders Programı — Dersliklere Göre"),
    ("lessons", "Ders Atama Listesi"),
    ("teachers", "Öğretmen Listesi"),
    ("classes", "Sınıf Listesi"),
    ("rooms", "Derslik Listesi"),
    ("subjects", "Ders Listesi"),
]


def build_sheet(data_store: dict, kind: str):
    """(sheet_title, headers, rows) for one export kind."""
    if kind == "timetable_classes":
        return "Sınıf Programı", *build_timetable_table(data_store, "classes")
    if kind == "timetable_teachers":
        return "Öğretmen Programı", *build_timetable_table(data_store, "teachers")
    if kind == "timetable_rooms":
        return "Derslik Programı", *build_timetable_table(data_store, "rooms")
    if kind == "lessons":
        return "Ders Atamaları", *build_lesson_list(data_store)
    if kind == "teachers":
        return "Öğretmenler", *build_entity_list(data_store, "ogretmenler")
    if kind == "classes":
        return "Sınıflar", *build_entity_list(data_store, "siniflar")
    if kind == "rooms":
        return "Derslikler", *build_entity_list(data_store, "derslikler")
    if kind == "subjects":
        return "Dersler", *build_entity_list(data_store, "dersler")
    raise ValueError(f"bilinmeyen dışa aktarım türü: {kind}")


def export_to_file(data_store: dict, path: str, kinds=None) -> str:
    """Writes the chosen sections to `path`; format follows the extension.

    CSV holds a single section (the format has no concept of sheets), so the first
    selected kind is written and the caller is expected to have said so.
    """
    kinds = list(kinds or ["timetable_classes"])
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        sheets = [build_sheet(data_store, k) for k in kinds]
        return write_xlsx(path, sheets)

    _title, headers, rows = build_sheet(data_store, kinds[0])
    return write_csv(path, headers, rows)
